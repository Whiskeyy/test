import tkinter as tk
from tkinter import messagebox, scrolledtext, ttk
import random
import time
import os
from datetime import datetime

# openpyxl optional
try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# -----------------------
# Konfiguration / Symbole
# -----------------------
SHAPES = [
    "Kreis_f", "Quadrat_f", "Raute_f", "Stern_f",
    "PfeilOben", "PfeilUnten", "PfeilLinks", "PfeilRechts",
    "Dreieck_f", "Doppelpfeil_H", "Doppelpfeil_V",
    "Kreis_h", "Quadrat_h", "Dreieck_h", "Stern_h", "Raute_h"
]

COLORS = [
    "red", "blue", "green", "orange",
    "purple", "brown", "cyan",
    "magenta", "gold", "darkgreen", "darkblue",
    "red", "blue", "gold", "magenta", "green"
]

TEST_SIZES = [3, 4, 5, 6, 7]
MEMORY_LIMIT_MS = 30000  # 30 Sekunden

EXCEL_FILENAME = "memorytest_results.xlsx"
SHEET_COLOR = "FARBIG"
SHEET_BW = "FARBLOS"
SHEET_QUESTION = "QUESTIONNAIRE"

GRAY_FILL = "#BBBBBB"
GRAY_OUTLINE = "#D3D3D3"
GRAY_STIPPLE = "gray50"

# UI styling constants
TITLE_FONT = ("Arial", 22, "bold")
SECTION_FONT = ("Arial", 14, "bold")
TEXT_FONT = ("Arial", 13)
SMALL_FONT = ("Arial", 11)
ACCENT = "#4B7CDA"  # dezentes blau


def safe_upper_alpha(s):
    return "".join(ch for ch in s if ch.isalpha()).upper()


# -----------------------
# App
# -----------------------
class MemoryTestApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Memory Test für Studie")
        self.root.minsize(950, 700)
        self.root.configure(bg="white")

        self.main = tk.Frame(root, padx=16, pady=16, bg="white")
        self.main.pack(fill="both", expand=True)

        # Input variables
        self.mother_var = tk.StringVar()
        self.father_var = tk.StringVar()
        self.birthyear_var = tk.StringVar()
        self.age_var = tk.StringVar()
        self.gender_var = tk.StringVar(value="M")

        # State
        self.test_variant = None
        self.current_test_index = 0
        self.sequence = []
        self.user_selections = []
        self.test_response_start = None
        self.memory_start = None
        self.memory_timer_id = None
        self.memory_progress_update_id = None
        self.merk_times = []
        self.response_times = []
        self.correct_counts = []

        # Canvases
        self.display_canvases = []
        self.input_canvases = []

        # Overlays: canvas -> overlay_id
        self.clicked_overlays = {}

        # Questionnaire
        self.q_vars = [tk.IntVar(value=0) for _ in range(7)]
        self.q8_text = tk.StringVar()

        # UI refs
        self.title_label = None
        self.status_label = None
        self.continue_button = None
        self.reset_button = None
        self.memory_progress = None
        self.memory_countdown_label = None

        self.build_start_menu()

    # -----------------------
    # Start Menu
    # -----------------------
    def build_start_menu(self):
        for w in self.main.winfo_children():
            w.destroy()

        tk.Label(self.main, text="Memory Test für Studie",
                 font=TITLE_FONT, bg="white", fg=ACCENT).pack(pady=(0, 12))

        tk.Label(self.main,
                 text="Bitte beantworten Sie folgende Fragen, damit wir eine anonyme Teilnehmer-ID bilden können.",
                 font=TEXT_FONT, bg="white").pack(pady=(0, 12))

        input_frame = tk.Frame(self.main, bg="white")
        input_frame.pack(pady=(0, 8))

        tk.Label(input_frame, text="Die ersten zwei Buchstaben des Vornamens Ihrer Mutter:",
                 bg="white").grid(row=0, column=0, sticky="e", padx=6, pady=6)
        tk.Entry(input_frame, textvariable=self.mother_var, width=25).grid(row=0, column=1, padx=6, pady=6)

        tk.Label(input_frame, text="Die letzten beiden Buchstaben des Vornamens Ihres Vaters:",
                 bg="white").grid(row=1, column=0, sticky="e", padx=6, pady=6)
        tk.Entry(input_frame, textvariable=self.father_var, width=25).grid(row=1, column=1, padx=6, pady=6)

        tk.Label(input_frame, text="Ihr Geburtsjahr:", bg="white").grid(row=2, column=0, sticky="e", padx=6, pady=6)
        tk.Entry(input_frame, textvariable=self.birthyear_var, width=12).grid(row=2, column=1, sticky="w", padx=6, pady=6)

        tk.Label(input_frame, text="Alter:", bg="white").grid(row=3, column=0, sticky="e", padx=6, pady=6)
        tk.Entry(input_frame, textvariable=self.age_var, width=6).grid(row=3, column=1, sticky="w", padx=6, pady=6)

        tk.Label(input_frame, text="Geschlecht:", bg="white").grid(row=4, column=0, sticky="e", padx=6, pady=6)
        gframe = tk.Frame(input_frame, bg="white")
        gframe.grid(row=4, column=1, sticky="w")
        tk.Radiobutton(gframe, text="M", variable=self.gender_var, value="M", bg="white").pack(side="left")
        tk.Radiobutton(gframe, text="W", variable=self.gender_var, value="W", bg="white").pack(side="left")
        tk.Radiobutton(gframe, text="D", variable=self.gender_var, value="D", bg="white").pack(side="left")

        if not OPENPYXL_AVAILABLE:
            tk.Label(self.main, text="Hinweis: openpyxl nicht installiert. Excel-Speicherung deaktiviert.",
                     fg="red", bg="white").pack(pady=(6, 4))

        tk.Button(self.main, text="Test starten", font=SECTION_FONT, bg=ACCENT, fg="white",
                  command=self.on_go_pressed).pack(pady=(10, 6))

    # -----------------------
    # TEXT FENSTER: Erklärung vor Teststart
    # -----------------------
    def on_go_pressed(self):
        for w in self.main.winfo_children():
            w.destroy()

        tk.Label(self.main, text="Memory Test für Studie",
                 font=TITLE_FONT, bg="white", fg=ACCENT).pack(pady=(0, 12))

        instructions = (
            "Sie erhalten gleich eine Reihe von Symbolen. Bitte merken Sie sich deren Reihenfolge. Sie haben für jede Merkphase 30 Sekunden Zeit.\n\n"
            #"\n"
            "Sie können die Merkphase vor Ablauf auch durch Klick auf\n"
            "„Weiter (Merkphase beenden)“ beenden, ansonsten endet sie automatisch nach 30 Sekunden.\n\n"
            "Nach der Merkphase klicken Sie die Symbole unterhalb in der Reihenfolge an, die Sie sich gemerkt haben.\n\n"
            "Viel Erfolg!"
        )
        tk.Label(self.main, text=instructions, font=TEXT_FONT, bg="white", justify=tk.CENTER).pack(pady=(0, 12))

        tk.Button(self.main, text="Weiter", font=SECTION_FONT, bg=ACCENT, fg="white",
                  command=self.on_start_pressed).pack(pady=(10, 6))

    # -----------------------
    # Start pressed: Validierung und Start
    # -----------------------
    def on_start_pressed(self):
        mother = safe_upper_alpha(self.mother_var.get().strip())
        father = safe_upper_alpha(self.father_var.get().strip())
        birth = self.birthyear_var.get().strip()
        age = self.age_var.get().strip()
        gender = self.gender_var.get().strip().upper()

        if len(mother) < 2:
            messagebox.showerror("Fehler", "Vorname der Mutter muss mindestens 2 Buchstaben enthalten.")
            return
        if len(father) < 2:
            messagebox.showerror("Fehler", "Vorname des Vaters muss mindestens 2 Buchstaben enthalten.")
            return
        if not (birth.isdigit() and len(birth) == 4):
            messagebox.showerror("Fehler", "Geburtsjahr muss 4-stellig sein (z.B. 1990).")
            return
        if not (age.isdigit() and 0 < int(age) < 130):
            messagebox.showerror("Fehler", "Bitte gültiges Alter eingeben.")
            return

        self.participant_id = f"{mother[:2]}{father[-2:]}{birth}"
        self.participant_age = int(age)
        self.participant_gender = gender

        self.test_variant = random.choice(["color", "bw"])
        self.current_test_index = 0
        self.merk_times = []
        self.response_times = []
        self.correct_counts = []

        for v in self.q_vars:
            v.set(0)
        self.q8_text.set("")

        self.clicked_overlays.clear()

        self.build_test_ui()
        self.root.after(200, self.start_test)

    # -----------------------
    # Test UI build
    # -----------------------
    def build_test_ui(self):
        for w in self.main.winfo_children():
            w.destroy()

        tk.Label(self.main, text="Memory Test für Studie",
                 font=TITLE_FONT, bg="white", fg=ACCENT).pack(pady=(0, 6))

        self.status_label = tk.Label(self.main, text="", font=SECTION_FONT, bg="white")
        self.status_label.pack(pady=(0, 8))

        # Display area (merkphase)
        self.display_area = tk.Frame(self.main, bg="white")
        self.display_area.pack(pady=(0, 8))

        # Memory progress and countdown (Variante B)
        progress_frame = tk.Frame(self.main, bg="white")
        progress_frame.pack(pady=(4, 6), fill="x", padx=8)

        self.memory_countdown_label = tk.Label(progress_frame, text="", font=TEXT_FONT, bg="white")
        self.memory_countdown_label.pack(side="left")

        self.memory_progress = ttk.Progressbar(progress_frame, orient="horizontal", length=400, mode="determinate")
        self.memory_progress.pack(side="right", padx=(8, 0))

        # Continue button (to end memory early)
        self.continue_button = tk.Button(self.main, text="Weiter (Merkphase beenden)",
                                         command=self.end_memory_phase, bg=ACCENT, fg="white")
        self.continue_button.pack_forget()

        # Input area (clickable items)
        self.input_area = tk.Frame(self.main, bg="white")
        self.input_area.pack(pady=(8, 8))

        self.display_canvases = []
        self.input_canvases = []

        for i in range(len(SHAPES)):
            c = tk.Canvas(self.display_area, width=88, height=88,
                          bg="white", highlightthickness=1)
            self.display_canvases.append(c)

        for i, shape in enumerate(SHAPES):
            c = tk.Canvas(self.input_area, width=68, height=68,
                          bg="white", highlightthickness=1)
            c.bind("<Button-1>", lambda e, s=shape: self.on_input_click(s))
            self.input_canvases.append(c)

        btn_frame = tk.Frame(self.main, bg="white")
        btn_frame.pack(pady=(6, 4))
        self.reset_button = tk.Button(btn_frame, text="Reset / Startmenü",
                                      command=self.reset_to_start, bg="white")
        self.reset_button.pack_forget()

    # -----------------------
    # Start a test
    # -----------------------
    def start_test(self):
        # Reset any existing progress updater
        if self.memory_progress_update_id:
            try:
                self.root.after_cancel(self.memory_progress_update_id)
            except Exception:
                pass
            self.memory_progress_update_id = None

        if self.current_test_index >= len(TEST_SIZES):
            self.after_tests_show_questionnaire()
            return

        self.clicked_overlays.clear()

        size = TEST_SIZES[self.current_test_index]
        self.sequence = random.sample(SHAPES, size)
        self.user_selections = []
        self.memory_start = time.time()
        self.memory_timer_id = None

        self.status_label.config(
            text=f"Test {self.current_test_index + 1}: Merke dir die Reihenfolge ({size} Symbole).",
            fg="black"
        )

        for c in self.display_canvases:
            c.grid_forget()

        for i in range(size):
            c = self.display_canvases[i]
            c.grid(row=0, column=i, padx=8, pady=8)
            shape = self.sequence[i]
            color = COLORS[SHAPES.index(shape)]
            draw_color = "black" if self.test_variant == "bw" else color
            self.draw_shape_display(c, shape, draw_color, gray=False)

        for c in self.input_canvases:
            c.grid_forget()

        # show continue button and start the timed progress
        self.continue_button.pack(pady=(6, 4))
        self.memory_progress['maximum'] = MEMORY_LIMIT_MS / 1000.0
        self.memory_progress['value'] = 0
        self._start_memory_progress_updater()
        self.memory_timer_id = self.root.after(MEMORY_LIMIT_MS, self.end_memory_phase)
        # Fortschrittsbalken sichtbar machen
        self.memory_progress.pack(side="right", padx=(8, 0))
        self.memory_countdown_label.pack(side="left")
        
    def _start_memory_progress_updater(self):
        # update every 100 ms for smoothness
        elapsed = time.time() - self.memory_start
        if elapsed < 0:
            elapsed = 0
        remaining = max(0, MEMORY_LIMIT_MS / 1000.0 - elapsed)
        # progress shows elapsed time (increasing), label shows remaining seconds
        self.memory_progress['value'] = min(self.memory_progress['maximum'], elapsed)
        self.memory_countdown_label.config(text=f"Verbleibende Zeit: {int(round(remaining))} s")
        if elapsed * 1000 >= MEMORY_LIMIT_MS:
            # stop updater; end_memory_phase will be called by timer
            self.memory_countdown_label.config(text=f"Verbleibende Zeit: 0 s")
            return
        self.memory_progress_update_id = self.root.after(100, self._start_memory_progress_updater)
        

    # -----------------------
    # End memory phase
    # -----------------------
    def end_memory_phase(self):
        # cancel scheduled events
        if self.memory_timer_id is not None:
            try:
                self.root.after_cancel(self.memory_timer_id)
            except Exception:
                pass
            self.memory_timer_id = None

        if self.memory_progress_update_id is not None:
            try:
                self.root.after_cancel(self.memory_progress_update_id)
            except Exception:
                pass
            self.memory_progress_update_id = None

        merk_time = round(
            min(time.time() - self.memory_start, MEMORY_LIMIT_MS / 1000.0), 3)
        self.merk_times.append(merk_time)
        # Fortschrittsbalken ausblenden
        self.memory_progress.pack_forget()
        self.memory_countdown_label.pack_forget()

        # hide continue button and reset progress UI
        self.continue_button.pack_forget()
        self.memory_progress['value'] = 0
        if self.memory_countdown_label:
            self.memory_countdown_label.config(text="")

        self.start_input_phase()

    # -----------------------
    # Input phase
    # -----------------------
    def start_input_phase(self):
        self.status_label.config(text="Eingabephase: Klicke die Symbole in der richtigen Reihenfolge.",
                                 fg="black")

        for c in self.display_canvases:
            c.grid_forget()

        self.clicked_overlays.clear()

        for i, c in enumerate(self.input_canvases):
            r = 0 if i < 8 else 1
            c.grid(row=r, column=i % 8, padx=6, pady=6)

            shape = SHAPES[i]
            color = COLORS[i]
            draw_color = "black" if self.test_variant == "bw" else color
            self.draw_shape_input(c, shape, draw_color, gray=False)

            c.bind("<Button-1>", lambda e, s=shape: self.on_input_click(s))

        self.test_response_start = time.time()
        self.user_selections = []

    # -----------------------
    # Input click
    # -----------------------
    def on_input_click(self, shape):
        try:
            idx = SHAPES.index(shape)
        except ValueError:
            return

        canvas = self.input_canvases[idx]

        if canvas in self.clicked_overlays:
            return

        if len(self.user_selections) >= len(self.sequence):
            return

        self.user_selections.append(shape)

        w_req = canvas.winfo_reqwidth()
        h_req = canvas.winfo_reqheight()

        overlay_id = canvas.create_rectangle(
            2, 2, w_req - 2, h_req - 2,
            fill=GRAY_FILL, outline=GRAY_OUTLINE, width=2,
            stipple=GRAY_STIPPLE
        )

        self.clicked_overlays[canvas] = overlay_id
        canvas.unbind("<Button-1>")

        if len(self.user_selections) == len(self.sequence):
            self.finish_test()

    # -----------------------
    # Finish test
    # -----------------------
    def finish_test(self):
        self.response_times.append(round(time.time() - self.test_response_start, 3))

        correct = sum(
            1 for a, b in zip(self.user_selections, self.sequence)
            if a == b
        )
        self.correct_counts.append(correct)

        self.current_test_index += 1

        if self.current_test_index < len(TEST_SIZES):
            self.root.after(700, self.start_test)
        else:
            if OPENPYXL_AVAILABLE:
                try:
                    self.save_test_results_to_excel()
                except Exception:
                    pass

            self.root.after(300, self.after_tests_show_questionnaire)

    # -----------------------
    # Draw helpers
    # -----------------------
    def draw_shape_display(self, canvas, shape, color, gray=False):
        canvas.delete("all")

        is_hollow = shape.endswith("_h")

        if gray:
            fill = GRAY_FILL
            outline = GRAY_OUTLINE
        else:
            if self.test_variant == "bw":
                fill = "black" if not is_hollow else "white"
                outline = "black"
            else:
                fill = color if not is_hollow else "white"
                outline = color

        width = 5 if is_hollow else 2
        self._draw_shape(canvas, shape, fill, outline, width)

    def draw_shape_input(self, canvas, shape, color, gray=False):
        canvas.delete("all")

        is_hollow = shape.endswith("_h")

        if gray:
            fill = GRAY_FILL
            outline = GRAY_OUTLINE
        else:
            if self.test_variant == "bw":
                fill = "black" if not is_hollow else "white"
                outline = "black"
            else:
                fill = color if not is_hollow else "white"
                outline = color

        width = 4 if is_hollow else 2
        self._draw_shape(canvas, shape, fill, outline, width)

    def _draw_shape(self, canvas, shape, fill, outline, width):
        canvas.delete("all")
        w = int(canvas.winfo_reqwidth())
        h = int(canvas.winfo_reqheight())
        s = min(w, h) / 100.0

        def sc(coords):
            return [coord * s for coord in coords]

        # --- Kreis ---
        if shape.startswith("Kreis"):
            canvas.create_oval(10*s, 10*s, 90*s, 90*s, fill=fill,
                               outline=outline, width=width)

        # --- Quadrat ---
        elif shape.startswith("Quadrat"):
            canvas.create_rectangle(12*s, 12*s, 88*s, 88*s,
                                    fill=fill, outline=outline, width=width)

        # --- Raute ---
        elif shape.startswith("Raute"):
            coords = sc([50, 8, 88, 50, 50, 92, 12, 50])
            canvas.create_polygon(coords, fill=fill,
                                  outline=outline, width=width)

        # --- Stern ---
        elif shape.startswith("Stern"):
            coords = sc([
                50, 12, 58, 34, 80, 34, 62, 50,
                68, 74, 50, 60, 32, 74, 38, 50,
                20, 34, 42, 34
            ])
            canvas.create_polygon(coords, fill=fill,
                                  outline=outline, width=width)

        # --- Dreieck ---
        elif shape.startswith("Dreieck"):
            coords = sc([50, 10, 90, 80, 10, 80])
            canvas.create_polygon(coords, fill=fill,
                                  outline=outline, width=width)

        # --- Pfeil Oben ---
        elif shape == "PfeilOben":
            coords = sc([
                50, 10, 78, 50, 62, 50,
                62, 90, 38, 90, 38, 50,
                22, 50
            ])
            canvas.create_polygon(coords, fill=fill,
                                  outline=outline, width=width)

        # --- Pfeil Unten ---
        elif shape == "PfeilUnten":
            coords = sc([50, 90, 78, 50, 62, 50, 62, 10, 38, 10, 38, 50, 22, 50])
            canvas.create_polygon(coords, fill=fill,
                                  outline=outline, width=width)

        # --- Pfeil Links ---
        elif shape == "PfeilLinks":
            coords = sc([10, 50, 50, 10, 50, 30, 90, 30,
                         90, 70, 50, 70, 50, 90])
            canvas.create_polygon(coords, fill=fill,
                                  outline=outline, width=width)

        # --- Pfeil Rechts ---
        elif shape == "PfeilRechts":
            coords = sc([90, 50, 50, 10, 50, 30, 10, 30,
                         10, 70, 50, 70, 50, 90])
            canvas.create_polygon(coords, fill=fill,
                                  outline=outline, width=width)

        # --- Doppelpfeil horizontal ---
        elif shape == "Doppelpfeil_H":
            canvas.create_line(18*s, 50*s, 82*s, 50*s,
                               fill=outline, width=width)
            canvas.create_polygon(sc([18, 50, 30, 40, 30, 60]),
                                  fill=outline, outline=outline)
            canvas.create_polygon(sc([82, 50, 70, 40, 70, 60]),
                                  fill=outline, outline=outline)

        # --- Doppelpfeil vertikal ---
        elif shape == "Doppelpfeil_V":
            canvas.create_line(50*s, 18*s, 50*s, 82*s,
                               fill=outline, width=width)
            canvas.create_polygon(sc([50, 18, 40, 30, 60, 30]),
                                  fill=outline, outline=outline)
            canvas.create_polygon(sc([50, 82, 40, 70, 60, 70]),
                                  fill=outline, outline=outline)

        # fallback
        else:
            canvas.create_oval(10*s, 10*s, 90*s, 90*s,
                               fill=fill, outline=outline, width=width)

    # -----------------------
    # After all tests: Questionnaire intro
    # -----------------------
    def after_tests_show_questionnaire(self):
        for w in self.main.winfo_children():
            w.destroy()

        tk.Label(self.main, text="Selbsteinschätzung (Fragebogen)",
                 font=TITLE_FONT, bg="white", fg=ACCENT).pack(pady=(0, 8))

        intro_text = (
            "Vielen Dank, Sie haben die Tests geschafft!\n\n"
            "Im folgenden Fragebogen schätzen Sie bitte Ihre eigene Erfahrung bei der Bearbeitung ein.\n"
            "Beantworten Sie bitte die Fragen ehrlich — es gibt kein richtig oder falsch.\n\n"
            "Klicken Sie auf „Zum Fragebogen“, um zu beginnen."
        )
        tk.Label(self.main, text=intro_text, font=TEXT_FONT, bg="white", justify=tk.CENTER).pack(pady=(0, 12))

        tk.Button(self.main, text="Zum Fragebogen", font=SECTION_FONT, bg=ACCENT, fg="white",
                  command=self.show_questionnaire_block1).pack(pady=(6, 4))

    # -----------------------
    # Questionnaire Block A
    # -----------------------
    def show_questionnaire_block1(self):
        for w in self.main.winfo_children():
            w.destroy()

        tk.Label(self.main, text="Selbsteinschätzung (Fragebogen)",
                 font=TITLE_FONT, bg="white", fg=ACCENT).pack(pady=(0, 8))

        self.q_frame = tk.Frame(self.main, bg="white")
        self.q_frame.pack(pady=(0, 6), padx=8, fill="both", expand=True)

        tk.Label(self.q_frame,
                 text="Teil A — Subjektive Performance",
                 font=SECTION_FONT, bg="white").pack(pady=(0, 6))

        questions = [
            "1. Das Bearbeiten des Tests fiel mir leicht",
            "2. Ich musste mich sehr konzentrieren um mir die Informationen zu merken.",
            "3. Ich hatte Schwierigkeiten mir mehrere Informationen gleichzeitig zu merken.",
            "4. Ich hatte Spaß an der Bearbeitung des Tests"
        ]

        for i, text in enumerate(questions):
            frame = tk.Frame(self.q_frame, bg="white")
            frame.pack(anchor="center", pady=10, fill="x")

            # Question label centered
            tk.Label(frame, text=text, anchor="center", bg="white", font=TEXT_FONT).pack(anchor="center", pady=(0, 8))
            
            # Radio buttons frame with labels on sides
            rb_main_frame = tk.Frame(frame, bg="white")
            rb_main_frame.pack(anchor="center", pady=4)
            
            # Left label
            tk.Label(rb_main_frame, text="stimme nicht zu", font=SMALL_FONT, bg="white").pack(side="left", padx=(0, 20))
            
            # Radio buttons centered
            rb_frame = tk.Frame(rb_main_frame, bg="white")
            rb_frame.pack(side="left")
            
            for val in range(1, 8):
                tk.Radiobutton(rb_frame, text=str(val),
                               variable=self.q_vars[i],
                               value=val, bg="white").pack(side="left", padx=8)
            
            # Right label
            tk.Label(rb_main_frame, text="stimme voll zu", font=SMALL_FONT, bg="white").pack(side="left", padx=(20, 0))

        btn_frame = tk.Frame(self.q_frame, bg="white")
        btn_frame.pack(pady=(12, 4))
        tk.Button(btn_frame, text="Weiter", command=self.show_questionnaire_block2, bg=ACCENT, fg="white").pack(side="left", padx=6)

    # -----------------------
    # Questionnaire Block B
    # -----------------------
    def show_questionnaire_block2(self):
        for w in self.q_frame.winfo_children():
            w.destroy()

        tk.Label(self.q_frame, text="Teil B — Weitere Einschätzung",
                 font=SECTION_FONT, bg="white").pack(pady=(0, 6))

        questions = [
            "5. Ich bin mit meiner Leistung im Test zufrieden",
            "6. Ich habe die korrekten Reihenfolgen reproduziert.",
            "7. Ich kann mir die angezeigte Reihenfolge für eine lange Zeit merken."
        ]

        for i, text in enumerate(questions, start=4):
            frame = tk.Frame(self.q_frame, bg="white")
            frame.pack(anchor="center", pady=10, fill="x")

            # Question label centered
            tk.Label(frame, text=text, anchor="center", bg="white", font=TEXT_FONT).pack(anchor="center", pady=(0, 8))
            
            # Radio buttons frame with labels on sides
            rb_main_frame = tk.Frame(frame, bg="white")
            rb_main_frame.pack(anchor="center", pady=4)
            
            # Left label
            tk.Label(rb_main_frame, text="stimme nicht zu", font=SMALL_FONT, bg="white").pack(side="left", padx=(0, 20))
            
            # Radio buttons centered
            rb_frame = tk.Frame(rb_main_frame, bg="white")
            rb_frame.pack(side="left")
            
            for val in range(1, 8):
                tk.Radiobutton(rb_frame, text=str(val),
                               variable=self.q_vars[i],
                               value=val, bg="white").pack(side="left", padx=8)
            
            # Right label
            tk.Label(rb_main_frame, text="stimme voll zu", font=SMALL_FONT, bg="white").pack(side="left", padx=(20, 0))

        tk.Label(self.q_frame,
                 text="8. Haben Sie eine Merkhilfe/Gedächtnisbrücke/Tricks verwendet?",
                 anchor="center", bg="white", font=TEXT_FONT).pack(anchor="w", pady=(8, 4))

        self.q8_textbox = scrolledtext.ScrolledText(self.q_frame,
                                                    width=80, height=5)
        self.q8_textbox.pack(pady=(0, 6))

        btn_frame = tk.Frame(self.q_frame, bg="white")
        btn_frame.pack(pady=(8, 4))

        tk.Button(btn_frame, text="Fragebogen abschicken", command=self.submit_questionnaire, bg=ACCENT, fg="white").pack(side="left", padx=6)
        tk.Button(btn_frame, text="Zurück", command=self.show_questionnaire_block1).pack(side="left", padx=6)

    # -----------------------
    # Submit questionnaire
    # -----------------------
    def submit_questionnaire(self):
        answers = [v.get() for v in self.q_vars]
        free_text = self.q8_textbox.get("1.0", "end").strip()
        self.q8_text.set(free_text)

        saved = False
        saved_path = None

        if OPENPYXL_AVAILABLE:
            try:
                saved, saved_path = self.save_questionnaire_to_excel(answers, free_text)
            except Exception as e:
                messagebox.showwarning("Warnung",
                                       f"Fehler beim Speichern des Fragebogens:\n{e}")
        else:
            # still proceed to show thanks page even without excel
            pass

        # Show final thank-you screen including results table
        self.show_thank_you_screen(saved, saved_path)

    # -----------------------
    # Thank you + Results screen
    # -----------------------
    def show_thank_you_screen(self, saved=False, saved_path=None):
        for w in self.main.winfo_children():
            w.destroy()

        tk.Label(self.main, text="Vielen Dank!", font=TITLE_FONT, bg="white", fg=ACCENT).pack(pady=(20, 6))
        tk.Label(self.main, text="Der Test ist beendet. Vielen Dank für Ihre Teilnahme.",
                 font=TEXT_FONT, bg="white").pack(pady=(0, 12))

        if saved and saved_path:
            tk.Label(self.main, text=f"Fragebogen gespeichert in: {saved_path}", font=SMALL_FONT, bg="white").pack()

        # Ergebnisse Tabelle
        table_frame = tk.Frame(self.main, bg="white", bd=1, relief="solid", padx=8, pady=8)
        table_frame.pack(pady=(16, 8))

        tk.Label(table_frame, text="Ergebnisse", font=SECTION_FONT, bg="white").grid(row=0, column=0, columnspan=len(TEST_SIZES), pady=(0, 8))

        # Header row: Test 1..N
        for i in range(len(TEST_SIZES)):
            tk.Label(table_frame, text=f"Test {i+1}", font=SMALL_FONT, bg="white", width=12, anchor="center", relief="ridge").grid(row=1, column=i, padx=2, pady=2)

        # Result row: "korrekt / size"
        for i in range(len(TEST_SIZES)):
            correct = self.correct_counts[i] if i < len(self.correct_counts) else 0
            size = TEST_SIZES[i]
            tk.Label(table_frame, text=f"{correct}/{size}", font=TEXT_FONT, bg="white", width=12, anchor="center", relief="ridge").grid(row=2, column=i, padx=2, pady=2)

        tk.Button(self.main, text="Zurück zum Start", command=self.reset_to_start, bg=ACCENT, fg="white").pack(pady=(12, 6))

    # -----------------------
    # Save Questionnaire
    # -----------------------
    def save_questionnaire_to_excel(self, answers, free_text):
        now_iso = datetime.now().isoformat(timespec="seconds")
        variant_label = "farbig" if self.test_variant == "color" else "farblos"

        row = [self.participant_id, variant_label] + answers + [free_text, now_iso]

        header = ["ID", "Testvariante"] + \
                 [f"Q{i+1}" for i in range(7)] + \
                 ["Q8_Freitext", "Datum"]

        base = next((d for d in [
            os.path.join(os.path.expanduser("~"), "Documents"),
            os.path.expanduser("~"), os.getcwd()
        ] if os.path.isdir(d)), os.getcwd())

        path = os.path.join(base, EXCEL_FILENAME)

        if os.path.exists(path):
            wb = load_workbook(path)
        else:
            wb = Workbook()

        if SHEET_QUESTION in wb.sheetnames:
            ws = wb[SHEET_QUESTION]
        else:
            ws = wb.create_sheet(SHEET_QUESTION)
            ws.append(header)

        ws.append(row)

        # remove empty default sheet
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            ws_def = wb["Sheet"]
            if ws_def.max_row == 1 and ws_def["A1"].value is None:
                wb.remove(ws_def)

        wb.save(path)
        return True, path

    # -----------------------
    # Save test results (once)
    # -----------------------
    def save_test_results_to_excel(self):
        now_iso = datetime.now().isoformat(timespec="seconds")
        variant_label = "farbig" if self.test_variant == "color" else "farblos"

        row = [
            self.participant_id,
            self.participant_age,
            self.participant_gender,
            variant_label,
        ]

        for i in range(len(TEST_SIZES)):
            row.extend([
                TEST_SIZES[i],
                self.merk_times[i] if i < len(self.merk_times) else "",
                self.response_times[i] if i < len(self.response_times) else "",
                self.correct_counts[i] if i < len(self.correct_counts) else ""
            ])

        header = [
            "ID", "Alter", "Geschlecht", "Testvariante"
        ]

        for size in TEST_SIZES:
            header += [f"{size}_Merkzeit", f"{size}_Reaktionszeit", f"{size}_Korrekt"]

        base = next((d for d in [
            os.path.join(os.path.expanduser("~"), "Documents"),
            os.path.expanduser("~"), os.getcwd()
        ] if os.path.isdir(d)), os.getcwd())

        path = os.path.join(base, EXCEL_FILENAME)

        if os.path.exists(path):
            wb = load_workbook(path)
        else:
            wb = Workbook()

        sheet_name = SHEET_COLOR if self.test_variant == "color" else SHEET_BW

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(header)

        ws.append(row)

        # Remove useless default sheet
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            ws_def = wb["Sheet"]
            if ws_def.max_row == 1 and ws_def["A1"].value is None:
                wb.remove(ws_def)

        wb.save(path)
        return True

    # -----------------------
    # Reset
    # -----------------------
    def reset_to_start(self):
        # cancel any scheduled callbacks to be safe
        try:
            if self.memory_timer_id:
                self.root.after_cancel(self.memory_timer_id)
        except Exception:
            pass
        try:
            if self.memory_progress_update_id:
                self.root.after_cancel(self.memory_progress_update_id)
        except Exception:
            pass

        self.build_start_menu()


# -----------------------
# Main
# -----------------------
if __name__ == "__main__":
    root = tk.Tk()
    style = ttk.Style(root)
    # use default theme but ensure progressbar looks decent
    style.theme_use('default')
    app = MemoryTestApp(root)
    root.mainloop()
