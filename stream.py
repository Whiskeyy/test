# app.py
import streamlit as st
import random
import time
import os
from datetime import datetime

# optional openpyxl
try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

# -----------------------
# Konfiguration / Symbole
# -----------------------
SHAPES = [
    "Kreis_f", "Quadrat_f", "Raute_f", "Stern_f",
    "PfeilOben", "PfeilUnten", "PfeilLinks", "PfeilRechts",
    "Dreieck_f", "Doppelpfeil_H", "Doppelpfeil_V",
    "Kreis_h", "Quadrat_h", "Dreieck_h", "Stern_h", "Raute_h"
]

COLORS = [
    "red", "blue", "green", "orange",
    "purple", "brown", "cyan",
    "magenta", "gold", "darkgreen", "darkblue",
    "red", "blue", "gold", "magenta", "green"
]

TEST_SIZES = [3, 4, 5, 6, 7]
MEMORY_LIMIT_MS = 30000  # 30 Sekunden

EXCEL_FILENAME = "memorytest_results.xlsx"
SHEET_COLOR = "FARBIG"
SHEET_BW = "FARBLOS"
SHEET_QUESTION = "QUESTIONNAIRE"

# small CSS to slightly center things
st.set_page_config(page_title="Memory Test für Studie", layout="wide")

# -----------------------
# Helpers
# -----------------------
def safe_upper_alpha(s):
    return "".join(ch for ch in s if ch.isalpha()).upper()

def svg_for_shape(shape, size_px=80, fill="white", stroke="black", stroke_width=2, gray=False):
    """
    Return SVG markup for a given shape name. Designed to resemble the original Tkinter polygons.
    If gray=True, render with gray fill/outline to indicate 'disabled' / clicked.
    """
    s = size_px
    viewbox = "0 0 100 100"
    # colors for gray state
    if gray:
        fill = "#BBBBBB"
        stroke = "#D3D3D3"
    is_hollow = shape.endswith("_h")

    # fill / stroke adjustments for hollow shapes
    if is_hollow:
        fill_color = "white" if not gray else fill
        stroke_color = stroke
    else:
        fill_color = fill
        stroke_color = stroke

    # polygons coordinates are scaled to 100x100 like in original
    def polygon(points):
        pts = " ".join([f"{x},{y}" for x, y in points])
        return f'<polygon points="{pts}" fill="{fill_color}" stroke="{stroke_color}" stroke-width="{stroke_width}" />'

    if shape.startswith("Kreis"):
        svg_body = f'<circle cx="50" cy="50" r="40" fill="{fill_color}" stroke="{stroke_color}" stroke-width="{stroke_width}" />'
    elif shape.startswith("Quadrat"):
        svg_body = f'<rect x="12" y="12" width="76" height="76" fill="{fill_color}" stroke="{stroke_color}" stroke-width="{stroke_width}" />'
    elif shape.startswith("Raute"):
        pts = [(50,8),(88,50),(50,92),(12,50)]
        svg_body = polygon(pts)
    elif shape.startswith("Stern"):
        pts = [
            (50,12),(58,34),(80,34),(62,50),
            (68,74),(50,60),(32,74),(38,50),
            (20,34),(42,34)
        ]
        svg_body = polygon(pts)
    elif shape.startswith("Dreieck"):
        pts = [(50,10),(90,80),(10,80)]
        svg_body = polygon(pts)
    elif shape == "PfeilOben":
        pts = [(50,10),(78,50),(62,50),(62,90),(38,90),(38,50),(22,50)]
        svg_body = polygon(pts)
    elif shape == "PfeilUnten":
        pts = [(50,90),(78,50),(62,50),(62,10),(38,10),(38,50),(22,50)]
        svg_body = polygon(pts)
    elif shape == "PfeilLinks":
        pts = [(10,50),(50,10),(50,30),(90,30),(90,70),(50,70),(50,90)]
        svg_body = polygon(pts)
    elif shape == "PfeilRechts":
        pts = [(90,50),(50,10),(50,30),(10,30),(10,70),(50,70),(50,90)]
        svg_body = polygon(pts)
    elif shape == "Doppelpfeil_H":
        line = f'<line x1="18" y1="50" x2="82" y2="50" stroke="{stroke_color}" stroke-width="{stroke_width}" />'
        left_tri = polygon([(18,50),(30,40),(30,60)])
        right_tri = polygon([(82,50),(70,40),(70,60)])
        svg_body = line + left_tri + right_tri
    elif shape == "Doppelpfeil_V":
        line = f'<line x1="50" y1="18" x2="50" y2="82" stroke="{stroke_color}" stroke-width="{stroke_width}" />'
        up_tri = polygon([(50,18),(40,30),(60,30)])
        down_tri = polygon([(50,82),(40,70),(60,70)])
        svg_body = line + up_tri + down_tri
    else:
        svg_body = f'<circle cx="50" cy="50" r="40" fill="{fill_color}" stroke="{stroke_color}" stroke-width="{stroke_width}" />'

    svg = f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="{viewbox}" width="{s}" height="{s}">{svg_body}</svg>'
    return svg

def show_svg(svg_markup):
    st.markdown(svg_markup, unsafe_allow_html=True)

def save_test_results_excel(session):
    """
    Save the test runs (merk_times, response_times, correct_counts) to an Excel file similar to original.
    Returns (saved_bool, path_or_error).
    """
    now_iso = datetime.now().isoformat(timespec="seconds")
    variant_label = "farbig" if session["test_variant"] == "color" else "farblos"

    row = [
        session["participant_id"],
        session.get("participant_age", ""),
        session.get("participant_gender", ""),
        variant_label,
    ]
    for i in range(len(TEST_SIZES)):
        row.extend([
            TEST_SIZES[i],
            session["merk_times"][i] if i < len(session["merk_times"]) else "",
            session["response_times"][i] if i < len(session["response_times"]) else "",
            session["correct_counts"][i] if i < len(session["correct_counts"]) else ""
        ])

    header = ["ID", "Alter", "Geschlecht", "Testvariante"]
    for size in TEST_SIZES:
        header += [f"{size}_Merkzeit", f"{size}_Reaktionszeit", f"{size}_Korrekt"]

    base = next((d for d in [
            os.path.join(os.path.expanduser("~"), "Documents"),
            os.path.expanduser("~"), os.getcwd()
        ] if os.path.isdir(d)), os.getcwd())

    path = os.path.join(base, EXCEL_FILENAME)
    try:
        if os.path.exists(path):
            wb = load_workbook(path)
        else:
            wb = Workbook()

        sheet_name = SHEET_COLOR if session["test_variant"] == "color" else SHEET_BW

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(header)

        ws.append(row)
        # cleanup default sheet
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            ws_def = wb["Sheet"]
            if ws_def.max_row == 1 and ws_def["A1"].value is None:
                wb.remove(ws_def)
        wb.save(path)
        return True, path
    except Exception as e:
        return False, str(e)

def save_questionnaire_excel(session, answers, free_text):
    now_iso = datetime.now().isoformat(timespec="seconds")
    variant_label = "farbig" if session["test_variant"] == "color" else "farblos"
    row = [session["participant_id"], variant_label] + answers + [free_text, now_iso]
    header = ["ID", "Testvariante"] + [f"Q{i+1}" for i in range(7)] + ["Q8_Freitext", "Datum"]

    base = next((d for d in [
            os.path.join(os.path.expanduser("~"), "Documents"),
            os.path.expanduser("~"), os.getcwd()
        ] if os.path.isdir(d)), os.getcwd())

    path = os.path.join(base, EXCEL_FILENAME)
    try:
        if os.path.exists(path):
            wb = load_workbook(path)
        else:
            wb = Workbook()

        if SHEET_QUESTION in wb.sheetnames:
            ws = wb[SHEET_QUESTION]
        else:
            ws = wb.create_sheet(SHEET_QUESTION)
            ws.append(header)

        ws.append(row)
        # cleanup default sheet
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            ws_def = wb["Sheet"]
            if ws_def.max_row == 1 and ws_def["A1"].value is None:
                wb.remove(ws_def)
        wb.save(path)
        return True, path
    except Exception as e:
        return False, str(e)


# -----------------------
# Session state init
# -----------------------
if "stage" not in st.session_state:
    st.session_state.stage = "start"  # stages: start, instructions, running_memory, input_phase, questionnaire_a, questionnaire_b, thanks
if "mother_var" not in st.session_state:
    st.session_state.mother_var = ""
if "father_var" not in st.session_state:
    st.session_state.father_var = ""
if "birthyear_var" not in st.session_state:
    st.session_state.birthyear_var = ""
if "age_var" not in st.session_state:
    st.session_state.age_var = ""
if "gender_var" not in st.session_state:
    st.session_state.gender_var = "M"

# test state
if "test_variant" not in st.session_state:
    st.session_state.test_variant = None
if "current_test_index" not in st.session_state:
    st.session_state.current_test_index = 0
if "sequence" not in st.session_state:
    st.session_state.sequence = []
if "user_selections" not in st.session_state:
    st.session_state.user_selections = []
if "merk_times" not in st.session_state:
    st.session_state.merk_times = []
if "response_times" not in st.session_state:
    st.session_state.response_times = []
if "correct_counts" not in st.session_state:
    st.session_state.correct_counts = []
if "clicked" not in st.session_state:
    st.session_state.clicked = {}  # shape -> True when clicked

# memory timer
if "memory_start" not in st.session_state:
    st.session_state.memory_start = None
if "memory_active" not in st.session_state:
    st.session_state.memory_active = False

# participant id etc
if "participant_id" not in st.session_state:
    st.session_state.participant_id = ""
if "participant_age" not in st.session_state:
    st.session_state.participant_age = None
if "participant_gender" not in st.session_state:
    st.session_state.participant_gender = ""

# questionnaire
if "q_vars" not in st.session_state:
    st.session_state.q_vars = [0]*7
if "q8_text" not in st.session_state:
    st.session_state.q8_text = ""

# small helpers to reset test progress
def reset_all():
    st.session_state.stage = "start"
    st.session_state.mother_var = ""
    st.session_state.father_var = ""
    st.session_state.birthyear_var = ""
    st.session_state.age_var = ""
    st.session_state.gender_var = "M"
    st.session_state.test_variant = None
    st.session_state.current_test_index = 0
    st.session_state.sequence = []
    st.session_state.user_selections = []
    st.session_state.merk_times = []
    st.session_state.response_times = []
    st.session_state.correct_counts = []
    st.session_state.clicked = {}
    st.session_state.memory_start = None
    st.session_state.memory_active = False
    st.session_state.participant_id = ""
    st.session_state.participant_age = None
    st.session_state.participant_gender = ""
    st.session_state.q_vars = [0]*7
    st.session_state.q8_text = ""

# -----------------------
# UI: Start Screen
# -----------------------
st.title("Memory Test für Studie")
if st.session_state.stage == "start":
    st.write("Bitte beantworten Sie folgende Fragen, damit wir eine anonyme Teilnehmer-ID bilden können.")
    col1, col2 = st.columns([2,3])
    with col1:
        st.session_state.mother_var = st.text_input("Die ersten zwei Buchstaben des Vornamens Ihrer Mutter:", value=st.session_state.mother_var)
        st.session_state.father_var = st.text_input("Die letzten beiden Buchstaben des Vornamens Ihres Vaters:", value=st.session_state.father_var)
    with col2:
        st.session_state.birthyear_var = st.text_input("Ihr Geburtsjahr:", value=st.session_state.birthyear_var, max_chars=4)
        st.session_state.age_var = st.text_input("Alter:", value=st.session_state.age_var, max_chars=3)
        st.session_state.gender_var = st.radio("Geschlecht:", options=["M","W","D"], index=["M","W","D"].index(st.session_state.gender_var if st.session_state.gender_var in ["M","W","D"] else "M"))
    if not OPENPYXL_AVAILABLE:
        st.warning("Hinweis: openpyxl nicht installiert. Excel-Speicherung deaktiviert.")
    if st.button("Test starten"):
        mother = safe_upper_alpha(st.session_state.mother_var.strip())
        father = safe_upper_alpha(st.session_state.father_var.strip())
        birth = st.session_state.birthyear_var.strip()
        age = st.session_state.age_var.strip()
        gender = st.session_state.gender_var.strip().upper()
        error = None
        if len(mother) < 2:
            error = "Vorname der Mutter muss mindestens 2 Buchstaben enthalten."
        elif len(father) < 2:
            error = "Vorname des Vaters muss mindestens 2 Buchstaben enthalten."
        elif not (birth.isdigit() and len(birth) == 4):
            error = "Geburtsjahr muss 4-stellig sein (z.B. 1990)."
        elif not (age.isdigit() and 0 < int(age) < 130):
            error = "Bitte gültiges Alter eingeben."
        if error:
            st.error(error)
        else:
            st.session_state.participant_id = f"{mother[:2]}{father[-2:]}{birth}"
            st.session_state.participant_age = int(age)
            st.session_state.participant_gender = gender
            st.session_state.test_variant = random.choice(["color", "bw"])
            st.session_state.current_test_index = 0
            st.session_state.merk_times = []
            st.session_state.response_times = []
            st.session_state.correct_counts = []
            st.session_state.q_vars = [0]*7
            st.session_state.q8_text = ""
            st.session_state.clicked = {}
            st.session_state.stage = "instructions"
            st.experimental_rerun()

# -----------------------
# Instructions
# -----------------------
if st.session_state.stage == "instructions":
    st.subheader("Instruktionen")
    st.write(
        "Sie erhalten gleich eine Reihe von Symbolen. Bitte merken Sie sich deren Reihenfolge. "
        "Sie haben für jede Merkphase 30 Sekunden Zeit.\n\n"
        "Sie können die Merkphase vor Ablauf auch durch Klick auf „Weiter (Merkphase beenden)“ beenden, "
        "ansonsten endet sie automatisch nach 30 Sekunden.\n\n"
        "Nach der Merkphase klicken Sie die Symbole unterhalb in der Reihenfolge an, die Sie sich gemerkt haben.\n\n"
        "Viel Erfolg!"
    )
    if st.button("Weiter"):
        st.session_state.stage = "running_memory"
        # reset clicked marks for new run
        st.session_state.clicked = {}
        st.session_state.user_selections = []
        st.session_state.sequence = []
        st.session_state.memory_start = None
        st.session_state.memory_active = False
        st.experimental_rerun()

# -----------------------
# Core: Start test / memory phase
# -----------------------
def start_test_round():
    size = TEST_SIZES[st.session_state.current_test_index]
    st.session_state.sequence = random.sample(SHAPES, size)
    st.session_state.user_selections = []
    st.session_state.clicked = {}
    st.session_state.memory_start = time.time()
    st.session_state.memory_active = True
    st.session_state.test_response_start = None

if st.session_state.stage == "running_memory":
    # if we arrived here without having generated a sequence yet, start one
    if not st.session_state.sequence:
        start_test_round()

    st.subheader(f"Test {st.session_state.current_test_index + 1}: Merke dir die Reihenfolge ({len(st.session_state.sequence)} Symbole).")

    # compute elapsed
    elapsed = 0.0
    if st.session_state.memory_start:
        elapsed = time.time() - st.session_state.memory_start
    remaining = max(0.0, MEMORY_LIMIT_MS/1000.0 - elapsed)
    progress = min(1.0, elapsed / (MEMORY_LIMIT_MS/1000.0))
    # progress bar
    st.progress(progress)
    st.write(f"Verbleibende Zeit: {int(round(remaining))} s")

    # show the sequence horizontally
    cols = st.columns(len(st.session_state.sequence))
    for i, shape in enumerate(st.session_state.sequence):
        color = COLORS[SHAPES.index(shape)]
        draw_color = "black" if st.session_state.test_variant == "bw" else color
        svg = svg_for_shape(shape, size_px=100, fill=draw_color if not shape.endswith("_h") else "white", stroke=draw_color, stroke_width=2, gray=False)
        with cols[i]:
            st.markdown(svg, unsafe_allow_html=True)

    # Button to end memory early
    if st.button("Weiter (Merkphase beenden)"):
        # record merk_time and go to input phase
        merk_time = round(min(time.time() - st.session_state.memory_start, MEMORY_LIMIT_MS / 1000.0), 3)
        st.session_state.merk_times.append(merk_time)
        st.session_state.memory_active = False
        st.session_state.stage = "input_phase"
        st.session_state.test_response_start = time.time()
        st.experimental_rerun()

    # Auto-finish when time exceeded
    if elapsed * 1000 >= MEMORY_LIMIT_MS:
        # record merk_time and go to input phase
        merk_time = round(min(elapsed, MEMORY_LIMIT_MS / 1000.0), 3)
        st.session_state.merk_times.append(merk_time)
        st.session_state.memory_active = False
        st.session_state.stage = "input_phase"
        st.session_state.test_response_start = time.time()
        st.experimental_rerun()

    # while memory active, auto-refresh the page every 700 ms so progress updates.
    # we inject a short meta-refresh only during memory phase.
    if st.session_state.memory_active:
        st.markdown(
            """
            <meta http-equiv="refresh" content="0.7">
            """, unsafe_allow_html=True
        )

# -----------------------
# Input phase: user clicks symbols in order
# -----------------------
if st.session_state.stage == "input_phase":
    st.subheader("Eingabephase: Klicke die Symbole in der richtigen Reihenfolge.")
    # grid of 16 shapes: show 8 on first row, 8 on second row
    grid_cols = st.columns(8)
    for i, shape in enumerate(SHAPES):
        col = grid_cols[i % 8]
        with col:
            clicked = st.session_state.clicked.get(shape, False)
            color = COLORS[i]
            draw_color = "black" if st.session_state.test_variant == "bw" else color
            svg = svg_for_shape(shape, size_px=80, fill=draw_color if not shape.endswith("_h") else "white", stroke=draw_color, stroke_width=2, gray=clicked)
            st.markdown(svg, unsafe_allow_html=True)
            # unique button key to avoid collisions
            btn_key = f"btn_{shape}_{st.session_state.current_test_index}"
            if not clicked:
                if st.button("Auswählen", key=btn_key):
                    # handle click
                    st.session_state.user_selections.append(shape)
                    st.session_state.clicked[shape] = True
                    # if finished selection for this round:
                    if len(st.session_state.user_selections) == len(st.session_state.sequence):
                        # record response time & correctness
                        resp_time = round(time.time() - st.session_state.test_response_start, 3)
                        st.session_state.response_times.append(resp_time)
                        correct = sum(1 for a,b in zip(st.session_state.user_selections, st.session_state.sequence) if a == b)
                        st.session_state.correct_counts.append(correct)
                        st.session_state.current_test_index += 1
                        # if more tests left: start next after short acknowledgement
                        if st.session_state.current_test_index < len(TEST_SIZES):
                            # prepare next test
                            st.success(f"Ergebnis Runde: {correct}/{len(st.session_state.sequence)} — Nächster Test beginnt gleich.")
                            # reset sequence to trigger new start in running_memory
                            st.session_state.sequence = []
                            st.session_state.user_selections = []
                            st.session_state.clicked = {}
                            st.session_state.stage = "running_memory"
                            st.session_state.memory_start = None
                            st.session_state.memory_active = False
                            st.experimental_rerun()
                        else:
                            # all tests done
                            if OPENPYXL_AVAILABLE:
                                try:
                                    save_test_results_excel(st.session_state)
                                except Exception:
                                    pass
                            st.session_state.stage = "questionnaire_a"
                            st.experimental_rerun()
                    else:
                        # continue; refresh to show overlay
                        st.experimental_rerun()
            else:
                st.button("Ausgewählt", disabled=True, key=f"sel_{shape}_{st.session_state.current_test_index}")

    # If user wants to reset mid-run
    if st.button("Reset / Startmenü"):
        reset_all()
        st.experimental_rerun()

# -----------------------
# Questionnaire A
# -----------------------
if st.session_state.stage == "questionnaire_a":
    st.header("Selbsteinschätzung (Fragebogen) — Teil A")
    st.write("Beantworte die folgenden Aussagen mit 1 (stimme nicht zu) bis 7 (stimme voll zu).")
    questions_a = [
        "1. Das Bearbeiten des Tests fiel mir leicht",
        "2. Ich musste mich sehr konzentrieren um mir die Informationen zu merken.",
        "3. Ich hatte Schwierigkeiten mir mehrere Informationen gleichzeitig zu merken.",
        "4. Ich hatte Spaß an der Bearbeitung des Tests"
    ]
    for i, q in enumerate(questions_a):
        st.write(q)
        st.session_state.q_vars[i] = st.radio("", options=[1,2,3,4,5,6,7], index=st.session_state.q_vars[i]-1 if (st.session_state.q_vars[i] and 1<=st.session_state.q_vars[i]<=7) else 3, key=f"qa{i}")

    if st.button("Weiter"):
        st.session_state.stage = "questionnaire_b"
        st.experimental_rerun()

# -----------------------
# Questionnaire B
# -----------------------
if st.session_state.stage == "questionnaire_b":
    st.header("Teil B — Weitere Einschätzung")
    questions_b = [
        "5. Ich bin mit meiner Leistung im Test zufrieden",
        "6. Ich habe die korrekten Reihenfolgen reproduziert.",
        "7. Ich kann mir die angezeigte Reihenfolge für eine lange Zeit merken."
    ]
    for j, q in enumerate(questions_b, start=4):
        st.write(q)
        st.session_state.q_vars[j] = st.radio("", options=[1,2,3,4,5,6,7], index=st.session_state.q_vars[j]-1 if (st.session_state.q_vars[j] and 1<=st.session_state.q_vars[j]<=7) else 3, key=f"qb{j}")

    st.write("8. Haben Sie eine Merkhilfe/Gedächtnisbrücke/Tricks verwendet?")
    st.session_state.q8_text = st.text_area("", value=st.session_state.q8_text, height=120)

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Fragebogen abschicken"):
            answers = st.session_state.q_vars.copy()
            free_text = st.session_state.q8_text.strip()
            saved = False
            saved_path = None
            if OPENPYXL_AVAILABLE:
                saved, saved_path = save_questionnaire_excel(st.session_state, answers, free_text)
            st.session_state.stage = "thanks"
            st.session_state.saved = saved
            st.session_state.saved_path = saved_path
            st.experimental_rerun()
    with col2:
        if st.button("Zurück"):
            st.session_state.stage = "questionnaire_a"
            st.experimental_rerun()

# -----------------------
# Thanks / Results
# -----------------------
if st.session_state.stage == "thanks":
    st.header("Vielen Dank!")
    st.write("Der Test ist beendet. Vielen Dank für Ihre Teilnahme.")
    if getattr(st.session_state, "saved", False) and st.session_state.saved_path:
        st.write(f"Fragebogen gespeichert in: {st.session_state.saved_path}")

    st.subheader("Ergebnisse")
    cols = st.columns(len(TEST_SIZES))
    for i in range(len(TEST_SIZES)):
        with cols[i]:
            correct = st.session_state.correct_counts[i] if i < len(st.session_state.correct_counts) else 0
            size = TEST_SIZES[i]
            st.write(f"Test {i+1}")
            st.write(f"{correct}/{size}")

    if st.button("Zurück zum Start"):
        reset_all()
        st.experimental_rerun()
